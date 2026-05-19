from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, CreateView, DetailView, DeleteView, TemplateView, FormView
from django.views.generic.base import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Exists, OuterRef, Q
from django import forms
import json as json_lib
from decimal import Decimal
from core.views import ClienteScopeMixin, PermissionRequiredMixin
from .models import Factura, CajaApertura, CajaMovimiento
from core.models import ConfigRestaurante
from comandas.models import Comanda
from inventario.models import Compra, CuentaPorPagar
from productos.models import Producto


class FacturaListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Factura
    template_name = "facturacion/list.html"
    context_object_name = "facturas"
    paginate_by = 20
    permission = "can_view_facturacion"

    def get_queryset(self):
        qs = super().get_queryset().select_related("comanda", "usuario")
        estado = self.request.GET.get("estado")
        if estado:
            qs = qs.filter(estado=estado)

        fecha_desde = self.request.GET.get("fecha_desde")
        fecha_hasta = self.request.GET.get("fecha_hasta")
        if fecha_desde:
            try:
                start = timezone.make_aware(timezone.datetime.strptime(fecha_desde, "%Y-%m-%d"))
                qs = qs.filter(fecha_emision__gte=start)
            except (ValueError, OverflowError):
                pass
        if fecha_hasta:
            try:
                end = timezone.make_aware(timezone.datetime.strptime(fecha_hasta, "%Y-%m-%d") + timezone.timedelta(days=1) - timezone.timedelta(seconds=1))
                qs = qs.filter(fecha_emision__lte=end)
            except (ValueError, OverflowError):
                pass

        search = self.request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(folio__icontains=search) |
                Q(cliente_nombre__icontains=search) |
                Q(comanda__codigo__icontains=search)
            )

        orden = self.request.GET.get("orden", "-fecha_emision")
        if orden in ["fecha_emision", "-fecha_emision", "total_con_impuestos", "-total_con_impuestos", "folio", "-folio"]:
            qs = qs.order_by(orden)
        else:
            qs = qs.order_by("-fecha_emision")

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_cliente()
        qs_base = Factura.objects.all()
        if cliente:
            qs_base = qs_base.filter(cliente=cliente)
        context["estados"] = Factura.Estado.choices
        context["estado_actual"] = self.request.GET.get("estado", "")
        context["fecha_desde"] = self.request.GET.get("fecha_desde", "")
        context["fecha_hasta"] = self.request.GET.get("fecha_hasta", "")
        context["search"] = self.request.GET.get("search", "")
        context["orden_actual"] = self.request.GET.get("orden", "-fecha_emision")
        context["total_ventas"] = qs_base.filter(estado=Factura.Estado.PAGADA).aggregate(
            total=Sum("total_con_impuestos")
        )["total"] or 0
        return context


class FacturaForm(forms.ModelForm):
    aplicar_iva = forms.BooleanField(label="Aplicar IVA", initial=True, required=False)
    aplicar_servicio = forms.BooleanField(label="Aplicar servicio", initial=True, required=False)
    monto_recibido = forms.DecimalField(label="Monto recibido", max_digits=10, decimal_places=2, required=False, help_text="Con cuanto paga el cliente (efectivo)")
    divisa_nombre = forms.CharField(label="Moneda alternativa", max_length=10, required=False, help_text="Ej: USD")
    divisa_monto = forms.DecimalField(label="Equivalente", max_digits=10, decimal_places=2, required=False)
    divisa_tasa = forms.DecimalField(label="Tasa de cambio", max_digits=10, decimal_places=4, required=False)

    class Meta:
        model = Factura
        fields = ["comanda", "cliente_nombre", "cliente_rfc", "metodo_pago", "descuento", "notas"]


@login_required
def comanda_data(request, pk):
    from comandas.models import Comanda
    if not request.user.has_perm("can_create_facturas"):
        return JsonResponse({"error": "Sin permiso"}, status=403)
    cliente = getattr(request.user, "cliente", None)
    qs = Comanda.objects.all()
    if cliente:
        qs = qs.filter(cliente=cliente)
    comanda = qs.filter(pk=pk).select_related("mesa").prefetch_related("items__producto").first()
    if not comanda:
        return JsonResponse({"error": "Comanda no encontrada"}, status=404)
    items = []
    for item in comanda.items.all():
        if not item.cancelado:
            items.append({
                "producto": item.producto.nombre,
                "cantidad": item.cantidad,
                "precio": float(item.precio_unitario),
                "subtotal": float(item.subtotal),
                "notas": item.notas,
            })
    subtotal = float(comanda.total)
    data = {
        "codigo": comanda.codigo,
        "mesa": comanda.mesa.numero,
        "items": items,
        "subtotal": subtotal,
        "total": subtotal,
    }
    return JsonResponse(data)

class FacturaCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Factura
    form_class = FacturaForm
    template_name = "facturacion/crear.html"
    success_url = reverse_lazy("facturacion:list")
    permission = "can_create_facturas"

    def get_initial(self):
        initial = super().get_initial()
        comanda_id = self.request.GET.get("comanda")
        if comanda_id:
            initial["comanda"] = comanda_id
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        comandas_con_factura = Factura.objects.filter(comanda=OuterRef("pk"))
        cliente = self.get_cliente()
        comandas_qs = Comanda.objects.all()
        if cliente:
            comandas_qs = comandas_qs.filter(cliente=cliente)
        form.fields["comanda"].queryset = comandas_qs.filter(
            estado__in=[Comanda.Estado.ABIERTA, Comanda.Estado.EN_COCINA, Comanda.Estado.EN_PREPARACION,
                        Comanda.Estado.LISTA, Comanda.Estado.SERVING, Comanda.Estado.CERRADA]
        ).exclude(Exists(comandas_con_factura)).order_by("-fecha_creacion")
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_cliente()
        config = ConfigRestaurante.get_config(cliente)
        context["tasa_iva"] = Decimal(str(config.tasa_impuesto))
        context["tasa_servicio"] = Decimal(str(config.porcentaje_servicio))
        context["initial_comanda_json"] = None
        comanda_id = self.get_initial().get("comanda")
        if comanda_id:
            from comandas.models import Comanda
            qs = Comanda.objects.all()
            if cliente:
                qs = qs.filter(cliente=cliente)
            c = qs.filter(pk=comanda_id).select_related("mesa").prefetch_related("items__producto").first()
            if c:
                items = []
                for item in c.items.all():
                    if not item.cancelado:
                        items.append({
                            "producto": item.producto.nombre,
                            "cantidad": item.cantidad,
                            "precio": float(item.precio_unitario),
                            "subtotal": float(item.subtotal),
                            "notas": item.notas,
                        })
                from json import dumps
                context["initial_comanda_json"] = dumps({
                    "codigo": c.codigo,
                    "mesa": c.mesa.numero,
                    "items": items,
                    "subtotal": float(c.total),
                    "total": float(c.total),
                })
        return context

    def form_valid(self, form):
        comanda = form.cleaned_data["comanda"]
        aplicar_iva = form.cleaned_data.get("aplicar_iva", False)
        aplicar_servicio = form.cleaned_data.get("aplicar_servicio", False)
        descuento = form.cleaned_data.get("descuento", 0)

        cliente_obj = comanda.cliente
        config = ConfigRestaurante.get_config(cliente_obj)
        tasa_iva = Decimal(str(config.tasa_impuesto))
        tasa_servicio = Decimal(str(config.porcentaje_servicio))

        subtotal = comanda.total
        iva = subtotal * tasa_iva if aplicar_iva else Decimal(0)
        servicio = subtotal * tasa_servicio if aplicar_servicio else Decimal(0)

        form.instance.cliente = cliente_obj
        form.instance.subtotal = subtotal
        form.instance.impuestos = iva
        form.instance.total_sin_impuestos = subtotal - descuento
        form.instance.propina = servicio
        form.instance.descuento = descuento
        form.instance.total_con_impuestos = subtotal + iva + servicio - descuento
        form.instance.usuario = self.request.user
        form.instance.estado = Factura.Estado.PAGADA
        form.instance.fecha_pago = timezone.now()

        monto_recibido = form.cleaned_data.get("monto_recibido")
        if monto_recibido:
            form.instance.monto_recibido = monto_recibido
            form.instance.cambio = monto_recibido - form.instance.total_con_impuestos

        divisa_nombre = form.cleaned_data.get("divisa_nombre")
        if divisa_nombre:
            form.instance.divisa_nombre = divisa_nombre
            form.instance.divisa_monto = form.cleaned_data.get("divisa_monto")
            form.instance.divisa_tasa = form.cleaned_data.get("divisa_tasa")

        response = super().form_valid(form)
        messages.success(self.request, f"Factura {form.instance.folio} generada")
        return response


class FacturaLlevarCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, View):
    template_name = "facturacion/crear_llevar.html"
    permission = "can_create_facturas"

    def get_cliente(self):
        return ClienteScopeMixin.get_cliente_static(self.request)

    def get(self, request):
        cliente = self.get_cliente()
        productos = Producto.objects.filter(cliente=cliente, activo=True).order_by("nombre") if cliente else []
        config = ConfigRestaurante.get_config(cliente)
        return render(request, self.template_name, {
            "productos": productos,
            "tasa_iva": config.tasa_impuesto if config else 0.15,
            "tasa_servicio": config.porcentaje_servicio if config else 0.10,
        })

    def post(self, request):
        cliente = self.get_cliente()
        if not cliente:
            messages.error(request, "Cliente no encontrado")
            return redirect("facturacion:list")

        productos_ids = request.POST.getlist("producto_id[]")
        cantidades = request.POST.getlist("cantidad[]")
        precios = request.POST.getlist("precio[]")

        items = []
        subtotal = Decimal(0)
        for pid, qty, price in zip(productos_ids, cantidades, precios):
            if not pid or not qty or Decimal(qty) <= 0:
                continue
            try:
                producto = Producto.objects.get(pk=pid, cliente=cliente)
                qty_dec = Decimal(qty)
                price_dec = Decimal(price) if price else producto.precio
                item_subtotal = qty_dec * price_dec
                items.append({
                    "producto_id": producto.pk,
                    "producto_nombre": producto.nombre,
                    "cantidad": int(qty_dec),
                    "precio_unitario": float(price_dec),
                    "subtotal": float(item_subtotal),
                })
                subtotal += item_subtotal
            except Producto.DoesNotExist:
                continue

        if not items:
            messages.error(request, "Agrega al menos un producto")
            return redirect("facturacion:crear_llevar")

        aplicar_iva = request.POST.get("aplicar_iva") == "on"
        aplicar_servicio = request.POST.get("aplicar_servicio") == "on"
        descuento = Decimal(request.POST.get("descuento", "0") or "0")
        metodo_pago = request.POST.get("metodo_pago", "efectivo")
        cliente_nombre = request.POST.get("cliente_nombre", "Consumidor Final")
        cliente_rfc = request.POST.get("cliente_rfc", "")
        monto_recibido = request.POST.get("monto_recibido", "")

        config = ConfigRestaurante.get_config(cliente)
        tasa_iva = Decimal(str(config.tasa_impuesto)) if config else Decimal("0.15")
        tasa_servicio = Decimal(str(config.porcentaje_servicio)) if config else Decimal("0.10")

        iva = subtotal * tasa_iva if aplicar_iva else Decimal(0)
        servicio = subtotal * tasa_servicio if aplicar_servicio else Decimal(0)
        total = subtotal + iva + servicio - descuento

        factura = Factura.objects.create(
            cliente=cliente,
            tipo=Factura.Tipo.LLEVAR,
            comanda=None,
            items_json=json_lib.dumps(items),
            cliente_nombre=cliente_nombre,
            cliente_rfc=cliente_rfc,
            subtotal=subtotal,
            impuestos=iva,
            descuento=descuento,
            propina=servicio,
            total_sin_impuestos=subtotal - descuento,
            total_con_impuestos=total,
            estado=Factura.Estado.PAGADA,
            metodo_pago=metodo_pago,
            fecha_pago=timezone.now(),
            usuario=request.user,
        )

        if metodo_pago == "efectivo" and monto_recibido:
            try:
                monto_dec = Decimal(monto_recibido)
                factura.monto_recibido = monto_dec
                factura.cambio = monto_dec - total
                factura.save(update_fields=["monto_recibido", "cambio"])
            except Exception:
                pass

        messages.success(request, f"Venta para llevar {factura.folio} registrada")
        return redirect("facturacion:imprimir", pk=factura.pk)


class FacturaDetailView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Factura
    template_name = "facturacion/detalle.html"
    context_object_name = "factura"
    permission = "can_view_facturacion"


class FacturaDeleteView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Factura
    success_url = reverse_lazy("facturacion:list")
    permission = "can_cancel_facturas"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f"Factura {self.object.folio} eliminada")
        return super().post(request, *args, **kwargs)


@login_required
def imprimir_factura(request, pk):
    cliente = getattr(request.user, 'cliente', None)
    qs = Factura.objects.all()
    if cliente:
        qs = qs.filter(cliente=cliente)
    factura = get_object_or_404(qs, pk=pk)
    return render(request, "facturacion/imprimir.html", {"factura": factura})


@login_required
def factura_escpos(request, pk):
    try:
        from core.escpos_utils import build_factura
        from core.models import ConfigRestaurante
        cliente = getattr(request.user, 'cliente', None)
        qs = Factura.objects.all()
        if cliente:
            qs = qs.filter(cliente=cliente)
        factura = get_object_or_404(qs, pk=pk)
        config = ConfigRestaurante.get_config(cliente)
        cols = int(request.GET.get("cols", "32"))
        data = build_factura(factura, config, cols=cols)
        return HttpResponse(data, content_type="application/octet-stream")
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return HttpResponse(f"Error ESC/POS: {e}\n\n{tb}", content_type="text/plain", status=500)


@login_required
def factura_escpos_tcp(request, pk):
    """Print directly to a network printer via TCP (port 9100)."""
    from core.escpos_utils import build_factura, send_tcp
    from core.models import ConfigRestaurante
    host = request.GET.get("host", "")
    port = int(request.GET.get("port", "9100"))
    if not host:
        return HttpResponse("Falta parametro host", status=400)
    cliente = getattr(request.user, 'cliente', None)
    qs = Factura.objects.all()
    if cliente:
        qs = qs.filter(cliente=cliente)
    factura = get_object_or_404(qs, pk=pk)
    config = ConfigRestaurante.get_config(cliente)
    data = build_factura(factura, config)
    try:
        send_tcp(data, host, port)
        return HttpResponse(f"Impreso en {host}:{port}")
    except Exception as e:
        return HttpResponse(f"Error al imprimir: {e}", status=500)


@login_required
def cierre_escpos(request):
    try:
        from core.escpos_utils import build_cierre
        from core.models import ConfigRestaurante
        cliente = getattr(request.user, 'cliente', None)
        config = ConfigRestaurante.get_config(cliente)
        view = CierreTicketView()
        view.request = request
        data = view.get_context_data()
        data['user'] = request.user
        cols = int(request.GET.get("cols", "32"))
        raw = build_cierre(data, config, cols=cols)
        return HttpResponse(raw, content_type="application/octet-stream")
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return HttpResponse(f"Error ESC/POS: {e}\n\n{tb}", content_type="text/plain", status=500)


@login_required
def cierre_escpos_tcp(request):
    from core.escpos_utils import build_cierre, send_tcp
    from core.models import ConfigRestaurante
    host = request.GET.get("host", "")
    port = int(request.GET.get("port", "9100"))
    if not host:
        return HttpResponse("Falta parametro host", status=400)
    cliente = getattr(request.user, 'cliente', None)
    config = ConfigRestaurante.get_config(cliente)
    view = CierreTicketView()
    view.request = request
    data = view.get_context_data()
    data['user'] = request.user
    raw = build_cierre(data, config)
    try:
        send_tcp(raw, host, port)
        return HttpResponse(f"Impreso en {host}:{port}")
    except Exception as e:
        return HttpResponse(f"Error al imprimir: {e}", status=500)


class AperturaCajaForm(forms.Form):
    monto_inicial = forms.DecimalField(
        label="Monto inicial en caja",
        min_value=0,
        max_digits=10,
        decimal_places=2,
        widget=forms.NumberInput(attrs={"step": "0.01", "placeholder": "0.00"})
    )
    notas = forms.CharField(
        label="Notas (opcional)",
        required=False,
        widget=forms.Textarea(attrs={"rows": 3, "placeholder": "Observaciones..."}),
    )


class AperturaCajaView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, FormView):
    template_name = "facturacion/apertura.html"
    form_class = AperturaCajaForm
    permission = "can_create_facturas"

    def dispatch(self, request, *args, **kwargs):
        cliente = self.get_cliente()
        caja_abierta = CajaApertura.objects.filter(
            estado=CajaApertura.Estado.ABIERTA,
            cliente=cliente
        ).first()
        if caja_abierta:
            messages.warning(request, "Ya hay una caja abierta. Debes cerrarla primero.")
            return redirect("facturacion:cierre_caja")
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["monto_inicial"].widget.attrs["value"] = "0.00"
        return form

    def form_valid(self, form):
        cliente = self.get_cliente()
        CajaApertura.objects.create(
            cliente=cliente,
            monto_inicial=form.cleaned_data["monto_inicial"],
            usuario_apertura=self.request.user,
            notas=form.cleaned_data.get("notas", ""),
        )
        messages.success(self.request, "Caja abierta correctamente")
        return redirect("facturacion:list")


class CierreCajaConfirmView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, FormView):
    template_name = "facturacion/cierre_confirm.html"
    permission = "can_create_facturas"
    DIFERENCIA_UMBRAL = 50

    def get_form_class(self):
        efectivo_esperado = getattr(self, "_efectivo_esperado", 0)

        class Form(forms.Form):
            monto_cierre_efectivo = forms.DecimalField(
                label="Efectivo contado al cierre",
                min_value=0,
                max_digits=10,
                decimal_places=2,
                widget=forms.NumberInput(attrs={"step": "0.01"})
            )
            notas_cierre = forms.CharField(
                label="Notas (obligatorio si hay faltante mayor a C$50)",
                required=False,
                widget=forms.Textarea(attrs={"rows": 3, "placeholder": "Explica la diferencia si aplica..."}),
            )

        return Form

    def get_efectivo_esperado(self):
        facturas_efectivo = Factura.objects.filter(
            fecha_emision__gte=self.caja.fecha_apertura,
            estado=Factura.Estado.PAGADA,
            metodo_pago=Factura.MetodoPago.EFECTIVO
        ).aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        return self.caja.monto_inicial + facturas_efectivo

    def dispatch(self, request, *args, **kwargs):
        cliente = self.get_cliente()
        self.caja = CajaApertura.objects.filter(
            estado=CajaApertura.Estado.ABIERTA,
            cliente=cliente
        ).first()
        if not self.caja:
            messages.warning(request, "No hay una caja abierta para cerrar.")
            return redirect("facturacion:apertura")
        self._efectivo_esperado = self.get_efectivo_esperado()
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["caja"] = self.caja
        context["diferencia_umbral"] = self.DIFERENCIA_UMBRAL
        context["efectivo_esperado"] = self._efectivo_esperado
        context["monto_inicial"] = self.caja.monto_inicial
        context["facturas_efectivo"] = self._efectivo_esperado - self.caja.monto_inicial
        return context

    def form_valid(self, form):
        monto_cierre = form.cleaned_data["monto_cierre_efectivo"]
        notas = form.cleaned_data.get("notas_cierre", "")
        diff = monto_cierre - self._efectivo_esperado

        if diff < -self.DIFERENCIA_UMBRAL and not notas.strip():
            form.add_error("notas_cierre", f"Debes explicar la diferencia de C${abs(diff):.2f}")
            return self.form_invalid(form)

        self.caja.notas = notas if notas else self.caja.notas
        self.caja.cerrar(monto_cierre, self.request.user)
        messages.success(self.request, f"Caja cerrada. Diferencia: C${self.caja.diferencia:+.2f}")
        return redirect(f"{reverse('facturacion:cierre_exitoso')}?caja_id={self.caja.pk}")


class CierreCajaView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, TemplateView):
    template_name = "facturacion/cierre_caja.html"
    permission = "can_view_facturacion"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fecha_str = self.request.GET.get("fecha")
        if fecha_str:
            try:
                from datetime import datetime
                hoy = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except (ValueError, OverflowError):
                hoy = timezone.localtime(timezone.now()).date()
        else:
            hoy = timezone.localtime(timezone.now()).date()

        start = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.min.time()))
        end = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.max.time()))

        cliente = self.get_cliente()
        facturas_qs = Factura.objects.all()
        comandas_qs = Comanda.objects.all()
        compras_qs = Compra.objects.all()
        cajas_qs = CajaApertura.objects.all()
        if cliente:
            facturas_qs = facturas_qs.filter(cliente=cliente)
            comandas_qs = comandas_qs.filter(cliente=cliente)
            compras_qs = compras_qs.filter(cliente=cliente)
            cajas_qs = cajas_qs.filter(cliente=cliente)

        caja_abierta = cajas_qs.filter(estado=CajaApertura.Estado.ABIERTA).first()
        context["caja_abierta"] = caja_abierta

        facturas_dia = facturas_qs.filter(fecha_emision__gte=start, fecha_emision__lte=end)

        context["hoy"] = hoy
        context["fecha_cierre"] = timezone.now()

        facturas_pagadas = facturas_dia.filter(estado=Factura.Estado.PAGADA)
        facturas_canceladas = facturas_dia.filter(estado=Factura.Estado.CANCELADA)
        facturas_pendientes = facturas_dia.filter(estado=Factura.Estado.PENDIENTE)

        context["total_facturas"] = facturas_dia.count()
        context["facturas_pagadas_count"] = facturas_pagadas.count()
        context["facturas_canceladas_count"] = facturas_canceladas.count()
        context["facturas_pendientes_count"] = facturas_pendientes.count()

        llevar_qs = facturas_pagadas.filter(tipo=Factura.Tipo.LLEVAR)
        context["llevar_count"] = llevar_qs.count()
        context["llevar_total"] = llevar_qs.aggregate(total=Sum("total_con_impuestos"))["total"] or 0

        total_ventas = facturas_pagadas.aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        total_iva = facturas_pagadas.aggregate(iva=Sum("impuestos"))["iva"] or 0
        total_servicio = facturas_pagadas.aggregate(servicio=Sum("propina"))["servicio"] or 0
        total_descuentos = facturas_pagadas.aggregate(descuento=Sum("descuento"))["descuento"] or 0
        context["total_ventas"] = total_ventas
        context["total_iva"] = total_iva
        context["total_servicio"] = total_servicio
        context["total_descuentos"] = total_descuentos
        context["total_neto"] = total_ventas - total_descuentos

        efectivo = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).aggregate(
            total=Sum("total_con_impuestos")
        )["total"] or 0
        tarjeta = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).aggregate(
            total=Sum("total_con_impuestos")
        )["total"] or 0
        transferencia = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).aggregate(
            total=Sum("total_con_impuestos")
        )["total"] or 0

        context["total_efectivo"] = efectivo
        context["total_tarjeta"] = tarjeta
        context["total_transferencia"] = transferencia

        ventas_por_metodo = [
            {"metodo": "Efectivo", "icono": "bi-cash-coin", "total": efectivo, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).count()},
            {"metodo": "Tarjeta", "icono": "bi-credit-card", "total": tarjeta, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).count()},
            {"metodo": "Transferencia", "icono": "bi-bank", "total": transferencia, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).count()},
        ]
        if total_ventas > 0:
            for item in ventas_por_metodo:
                item["percent"] = round((item["total"] / total_ventas) * 100)
        else:
            for item in ventas_por_metodo:
                item["percent"] = 0

        context["ventas_por_metodo"] = ventas_por_metodo

        context["facturas_detalladas"] = facturas_pagadas.select_related("comanda", "usuario").order_by("fecha_emision")
        context["facturas_canceladas_detalladas"] = facturas_canceladas.select_related("comanda", "usuario").order_by("fecha_emision")

        comandas_cerradas = comandas_qs.filter(fecha_cierre__gte=start, fecha_cierre__lte=end)
        context["comandas_cerradas_count"] = comandas_cerradas.count()
        context["fecha_seleccionada"] = hoy

        compras_dia = compras_qs.filter(fecha__gte=start, fecha__lte=end)
        compras_recibidas = compras_dia.filter(estado=Compra.Estado.RECIBIDA)
        compras_pendientes = compras_dia.filter(estado=Compra.Estado.PENDIENTE)
        compras_canceladas = compras_dia.filter(estado=Compra.Estado.CANCELADA)

        total_compras = sum(c.total for c in compras_recibidas)

        context["compras_recibidas_count"] = compras_recibidas.count()
        context["compras_pendientes_count"] = compras_pendientes.count()
        context["compras_canceladas_count"] = compras_canceladas.count()
        context["total_compras"] = total_compras
        context["compras_detalladas"] = compras_recibidas.prefetch_related("items").order_by("fecha")

        aperturas_dia = cajas_qs.filter(fecha_apertura__gte=start, fecha_apertura__lte=end).select_related("usuario_apertura")
        context["aperturas_dia"] = aperturas_dia

        movimientos = CajaMovimiento.objects.none()
        total_gastos = 0
        total_retiros = 0
        if caja_abierta:
            movimientos = caja_abierta.movimientos.select_related("usuario").order_by("-fecha")
            total_gastos = movimientos.filter(tipo=CajaMovimiento.Tipo.GASTO).aggregate(total=Sum("monto"))["total"] or 0
            total_retiros = movimientos.filter(tipo=CajaMovimiento.Tipo.RETIRO).aggregate(total=Sum("monto"))["total"] or 0
        else:
            cajas_ids = list(cajas_qs.filter(fecha_apertura__gte=start, fecha_apertura__lte=end).values_list("pk", flat=True))
            movimientos = CajaMovimiento.objects.filter(caja_id__in=cajas_ids).select_related("usuario").order_by("-fecha")
            totals = movimientos.aggregate(
                gastos=Sum("monto", filter=Q(tipo=CajaMovimiento.Tipo.GASTO)),
                retiros=Sum("monto", filter=Q(tipo=CajaMovimiento.Tipo.RETIRO)),
            )
            total_gastos = totals["gastos"] or 0
            total_retiros = totals["retiros"] or 0
        context["movimientos"] = movimientos
        context["total_gastos_mov"] = total_gastos
        context["total_retiros_mov"] = total_retiros

        balance_neto = total_ventas - total_compras - total_gastos - total_retiros
        context["balance_neto"] = balance_neto
        context["balance_color"] = "success" if balance_neto >= 0 else "danger"

        return context


class CierreCajaPdfView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, TemplateView):
    permission = "can_view_facturacion"

    def get(self, request, *args, **kwargs):
        from django.http import HttpResponse
        from core.pdf_utils import generate_cierre_pdf
        from core.models import ConfigRestaurante
        from datetime import datetime

        fecha_str = request.GET.get("fecha")
        if fecha_str:
            try:
                hoy = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except (ValueError, OverflowError):
                hoy = timezone.localtime(timezone.now()).date()
        else:
            hoy = timezone.localtime(timezone.now()).date()

        start = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.min.time()))
        end = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.max.time()))

        cliente = self.get_cliente()
        facturas_qs = Factura.objects.filter(cliente=cliente) if cliente else Factura.objects.all()
        compras_qs = Compra.objects.filter(cliente=cliente) if cliente else Compra.objects.all()
        cajas_qs = CajaApertura.objects.filter(cliente=cliente) if cliente else CajaApertura.objects.all()

        facturas_pagadas = facturas_qs.filter(fecha_emision__gte=start, fecha_emision__lte=end, estado=Factura.Estado.PAGADA)
        facturas_canceladas = facturas_qs.filter(fecha_emision__gte=start, fecha_emision__lte=end, estado=Factura.Estado.CANCELADA)
        compras_recibidas = compras_qs.filter(fecha__gte=start, fecha__lte=end, estado=Compra.Estado.RECIBIDA)

        llevar_qs = facturas_pagadas.filter(tipo=Factura.Tipo.LLEVAR)
        llevar_count = llevar_qs.count()
        llevar_total = llevar_qs.aggregate(total=Sum("total_con_impuestos"))["total"] or 0

        total_ventas = facturas_pagadas.aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        total_iva = facturas_pagadas.aggregate(iva=Sum("impuestos"))["iva"] or 0
        total_servicio = facturas_pagadas.aggregate(servicio=Sum("propina"))["servicio"] or 0
        total_descuentos = facturas_pagadas.aggregate(descuento=Sum("descuento"))["descuento"] or 0
        total_compras = sum(c.total for c in compras_recibidas)

        efectivo = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        tarjeta = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        transferencia = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).aggregate(total=Sum("total_con_impuestos"))["total"] or 0

        ventas_por_metodo = [
            {"metodo": "Efectivo", "total": efectivo, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).count()},
            {"metodo": "Tarjeta", "total": tarjeta, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).count()},
            {"metodo": "Transferencia", "total": transferencia, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).count()},
        ]

        aperturas_cerradas = cajas_qs.filter(fecha_apertura__gte=start, fecha_apertura__lte=end, estado=CajaApertura.Estado.CERRADA).first()
        config = ConfigRestaurante.get_config(cliente)

        caja_abierta = cajas_qs.filter(estado=CajaApertura.Estado.ABIERTA).first()
        movimientos = CajaMovimiento.objects.none()
        if caja_abierta:
            movimientos = caja_abierta.movimientos.all()
        cajas_cerradas_hoy = cajas_qs.filter(
            fecha_apertura__gte=start, fecha_apertura__lte=end,
            estado=CajaApertura.Estado.CERRADA
        )
        if cajas_cerradas_hoy.exists():
            movimientos = movimientos | CajaMovimiento.objects.filter(caja__in=cajas_cerradas_hoy)
        movimientos = movimientos.select_related("usuario").order_by("-fecha")
        total_gastos_mov = movimientos.filter(tipo=CajaMovimiento.Tipo.GASTO).aggregate(total=Sum("monto"))["total"] or 0
        total_retiros_mov = movimientos.filter(tipo=CajaMovimiento.Tipo.RETIRO).aggregate(total=Sum("monto"))["total"] or 0

        buffer = generate_cierre_pdf(
            aperturas_cerradas, facturas_pagadas, facturas_canceladas, compras_recibidas,
            ventas_por_metodo, total_ventas, total_compras, total_iva, total_servicio, total_descuentos,
            total_ventas - total_compras, config.nombre, config.simbolo_moneda, hoy,
            movimientos=movimientos, total_gastos_mov=total_gastos_mov, total_retiros_mov=total_retiros_mov,
            llevar_count=llevar_count, llevar_total=llevar_total
        )
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="cierre_{hoy.strftime("%Y%m%d")}.pdf"'
        return response


class CierreTicketView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, TemplateView):
    template_name = "facturacion/cierre_ticket.html"
    permission = "can_view_facturacion"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fecha_str = self.request.GET.get("fecha")
        if fecha_str:
            from datetime import datetime
            hoy = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        else:
            hoy = timezone.localtime(timezone.now()).date()

        start = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.min.time()))
        end = timezone.make_aware(timezone.datetime.combine(hoy, timezone.datetime.max.time()))

        cliente = self.get_cliente()
        facturas_qs = Factura.objects.all()
        comandas_qs = Comanda.objects.all()
        compras_qs = Compra.objects.all()
        cajas_qs = CajaApertura.objects.all()
        if cliente:
            facturas_qs = facturas_qs.filter(cliente=cliente)
            comandas_qs = comandas_qs.filter(cliente=cliente)
            compras_qs = compras_qs.filter(cliente=cliente)
            cajas_qs = cajas_qs.filter(cliente=cliente)

        context["caja_abierta"] = cajas_qs.filter(estado=CajaApertura.Estado.ABIERTA).first()
        facturas_dia = facturas_qs.filter(fecha_emision__gte=start, fecha_emision__lte=end)
        context["fecha_seleccionada"] = hoy
        context["fecha_cierre"] = timezone.now()

        facturas_pagadas = facturas_dia.filter(estado=Factura.Estado.PAGADA)
        facturas_canceladas = facturas_dia.filter(estado=Factura.Estado.CANCELADA)
        context["facturas_pagadas_count"] = facturas_pagadas.count()
        context["facturas_canceladas_count"] = facturas_canceladas.count()

        llevar_qs = facturas_pagadas.filter(tipo=Factura.Tipo.LLEVAR)
        context["llevar_count"] = llevar_qs.count()
        context["llevar_total"] = llevar_qs.aggregate(total=Sum("total_con_impuestos"))["total"] or 0

        total_ventas = facturas_pagadas.aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        total_iva = facturas_pagadas.aggregate(iva=Sum("impuestos"))["iva"] or 0
        total_servicio = facturas_pagadas.aggregate(servicio=Sum("propina"))["servicio"] or 0
        total_descuentos = facturas_pagadas.aggregate(descuento=Sum("descuento"))["descuento"] or 0
        context["total_ventas"] = total_ventas
        context["total_iva"] = total_iva
        context["total_servicio"] = total_servicio
        context["total_descuentos"] = total_descuentos

        efectivo = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        tarjeta = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).aggregate(total=Sum("total_con_impuestos"))["total"] or 0
        transferencia = facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).aggregate(total=Sum("total_con_impuestos"))["total"] or 0

        ventas_por_metodo = [
            {"metodo": "Efectivo", "total": efectivo, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.EFECTIVO).count()},
            {"metodo": "Tarjeta", "total": tarjeta, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TARJETA).count()},
            {"metodo": "Transferencia", "total": transferencia, "count": facturas_pagadas.filter(metodo_pago=Factura.MetodoPago.TRANSFERENCIA).count()},
        ]
        context["ventas_por_metodo"] = ventas_por_metodo

        comandas_cerradas = comandas_qs.filter(fecha_cierre__gte=start, fecha_cierre__lte=end)
        context["comandas_cerradas_count"] = comandas_cerradas.count()

        compras_dia = compras_qs.filter(fecha__gte=start, fecha__lte=end)
        compras_recibidas = compras_dia.filter(estado=Compra.Estado.RECIBIDA)
        compras_pendientes = compras_dia.filter(estado=Compra.Estado.PENDIENTE)
        compras_canceladas = compras_dia.filter(estado=Compra.Estado.CANCELADA)
        total_compras = sum(c.total for c in compras_recibidas)
        context["compras_recibidas_count"] = compras_recibidas.count()
        context["compras_pendientes_count"] = compras_pendientes.count()
        context["compras_canceladas_count"] = compras_canceladas.count()
        context["total_compras"] = total_compras
        context["compras_detalladas"] = compras_recibidas.prefetch_related("items").order_by("fecha")

        caja_abierta = cajas_qs.filter(estado=CajaApertura.Estado.ABIERTA).first()
        movimientos = CajaMovimiento.objects.none()
        total_gastos_mov = 0
        total_retiros_mov = 0
        if caja_abierta:
            movimientos = caja_abierta.movimientos.select_related("usuario").order_by("-fecha")
            total_gastos_mov = movimientos.filter(tipo=CajaMovimiento.Tipo.GASTO).aggregate(total=Sum("monto"))["total"] or 0
            total_retiros_mov = movimientos.filter(tipo=CajaMovimiento.Tipo.RETIRO).aggregate(total=Sum("monto"))["total"] or 0
        else:
            cajas_dia = cajas_qs.filter(fecha_apertura__gte=start, fecha_apertura__lte=end)
            for caja in cajas_dia:
                movs = caja.movimientos.select_related("usuario").all()
                movimientos = movimientos | movs
                total_gastos_mov += movs.filter(tipo=CajaMovimiento.Tipo.GASTO).aggregate(total=Sum("monto"))["total"] or 0
                total_retiros_mov += movs.filter(tipo=CajaMovimiento.Tipo.RETIRO).aggregate(total=Sum("monto"))["total"] or 0
            movimientos = movimientos.order_by("-fecha")
        context["movimientos"] = movimientos
        context["total_gastos_mov"] = total_gastos_mov
        context["total_retiros_mov"] = total_retiros_mov

        context["balance_neto"] = total_ventas - total_compras - total_gastos_mov - total_retiros_mov

        aperturas_dia = cajas_qs.filter(fecha_apertura__gte=start, fecha_apertura__lte=end).select_related("usuario_apertura")
        context["aperturas_dia"] = aperturas_dia
        return context


class HistorialCierresView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = CajaApertura
    template_name = "facturacion/historial_cierres.html"
    context_object_name = "cierres"
    paginate_by = 20
    permission = "can_view_facturacion"

    def get_queryset(self):
        qs = CajaApertura.objects.filter(estado=CajaApertura.Estado.CERRADA).select_related("usuario_apertura", "usuario_cierre")
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)

        fecha_desde = self.request.GET.get("fecha_desde")
        fecha_hasta = self.request.GET.get("fecha_hasta")
        usuario = self.request.GET.get("usuario")

        if fecha_desde:
            try:
                start = timezone.make_aware(timezone.datetime.strptime(fecha_desde, "%Y-%m-%d"))
                qs = qs.filter(fecha_apertura__gte=start)
            except (ValueError, OverflowError):
                pass
        if fecha_hasta:
            try:
                end = timezone.make_aware(timezone.datetime.strptime(fecha_hasta, "%Y-%m-%d") + timezone.timedelta(days=1) - timezone.timedelta(seconds=1))
                qs = qs.filter(fecha_apertura__lte=end)
            except (ValueError, OverflowError):
                pass
        if usuario:
            qs = qs.filter(usuario_apertura_id=usuario)

        return qs.order_by("-fecha_apertura")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["fecha_desde"] = self.request.GET.get("fecha_desde", "")
        context["fecha_hasta"] = self.request.GET.get("fecha_hasta", "")
        context["usuario_filtro"] = self.request.GET.get("usuario", "")

        cliente = self.get_cliente()
        qs = CajaApertura.objects.filter(estado=CajaApertura.Estado.CERRADA)
        if cliente:
            qs = qs.filter(cliente=cliente)
        context["usuarios"] = CajaApertura.objects.filter(cliente=cliente).values_list("usuario_apertura__id", "usuario_apertura__first_name", "usuario_apertura__last_name", "usuario_apertura__username").distinct()
        context["stats"] = {
            "total_cierres": qs.count(),
            "diferencia_promedio": qs.aggregate(avg=Sum("diferencia"))["avg"] or 0,
            "total_faltante": qs.filter(diferencia__lt=0).aggregate(total=Sum("diferencia"))["total"] or 0,
            "total_sobrante": qs.filter(diferencia__gt=0).aggregate(total=Sum("diferencia"))["total"] or 0,
        }
        return context


class CierreExitosoView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, TemplateView):
    template_name = "facturacion/cierre_exitoso.html"
    permission = "can_view_facturacion"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        caja_id = self.request.GET.get("caja_id")
        if caja_id:
            context["caja"] = CajaApertura.objects.filter(pk=caja_id).select_related("usuario_apertura", "usuario_cierre").first()
        return context


class CajaMovimientoCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, View):
    permission = "can_create_facturas"

    def post(self, request):
        caja_id = request.POST.get("caja_id")
        tipo = request.POST.get("tipo", CajaMovimiento.Tipo.GASTO)
        monto = request.POST.get("monto", "0")
        descripcion = request.POST.get("descripcion", "")

        if not descripcion:
            messages.error(request, "La descripcion es obligatoria")
            return redirect("facturacion:cierre_caja")

        from decimal import Decimal
        try:
            monto_dec = Decimal(monto)
        except Exception:
            messages.error(request, "Monto no valido")
            return redirect("facturacion:cierre_caja")

        if monto_dec <= 0:
            messages.error(request, "El monto debe ser mayor a cero")
            return redirect("facturacion:cierre_caja")

        caja = get_object_or_404(CajaApertura, pk=caja_id, estado=CajaApertura.Estado.ABIERTA)
        CajaMovimiento.objects.create(
            caja=caja,
            tipo=tipo,
            monto=monto_dec,
            descripcion=descripcion,
            usuario=request.user,
        )
        msg = "Gasto registrado" if tipo == CajaMovimiento.Tipo.GASTO else "Retiro registrado"
        messages.success(request, f"{msg}: {descripcion} ({monto_dec})")
        return redirect("facturacion:cierre_caja")


@login_required
def facturas_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Border, Side

    from core.views import ClienteScopeMixin

    cliente = ClienteScopeMixin.get_cliente_static(request)
    qs = Factura.objects.select_related("comanda__mesa")
    if cliente:
        qs = qs.filter(cliente=cliente)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    header_font = Font(name="Arial", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Folio", "Fecha", "Mesa", "Cliente", "Metodo Pago", "Subtotal", "IVA", "Servicio", "Descuento", "Total", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border

    for row, f in enumerate(qs, 2):
        data = [
            f.folio,
            timezone.localtime(f.fecha_emision).strftime("%d/%m/%Y %H:%M"),
            f.comanda.mesa.numero if f.comanda else "",
            f.cliente_nombre or "",
            f.get_metodo_pago_display(),
            float(f.subtotal),
            float(f.impuestos),
            float(f.propina),
            float(f.descuento),
            float(f.total_con_impuestos),
            f.get_estado_display(),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    for c in "FGHIJ": ws.column_dimensions[c].width = 14
    ws.column_dimensions["K"].width = 12

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="facturas.xlsx"'
    wb.save(response)
    return response
