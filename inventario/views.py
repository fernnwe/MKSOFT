from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import F, Sum, Q, ExpressionWrapper, DecimalField
from django.db import IntegrityError, models
from django.utils import timezone
from django import forms
from django.views.decorators.http import require_GET
from django.http import JsonResponse, HttpResponse
from core.views import ClienteScopeMixin, PermissionRequiredMixin
from .models import Ingrediente, Inventario, MovimientoInventario, Compra, CompraItem, CuentaPorPagar, Proveedor
from core.models import ConfigRestaurante


class InventarioListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Inventario
    template_name = "inventario/list.html"
    context_object_name = "inventario"
    paginate_by = 20
    permission = "can_view_inventario"

    def get_queryset(self):
        qs = super().get_queryset().select_related("ingrediente")
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(ingrediente__cliente=cliente)

        alertas = self.request.GET.get("alertas")
        if alertas == "1":
            qs = qs.filter(
                Q(cantidad_actual__lte=F("ingrediente__stock_minimo"), ingrediente__stock_minimo__gt=0) |
                Q(cantidad_actual=0)
            )

        estado = self.request.GET.get("estado")
        if estado == "bajo":
            qs = qs.filter(
                Q(cantidad_actual__lte=F("ingrediente__stock_minimo"), ingrediente__stock_minimo__gt=0) |
                Q(cantidad_actual=0)
            )
        elif estado == "ok":
            qs = qs.filter(
                cantidad_actual__gt=F("ingrediente__stock_minimo")
            )

        categoria = self.request.GET.get("categoria")
        if categoria:
            qs = qs.filter(ingrediente__categoria=categoria)

        search = self.request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(ingrediente__nombre__icontains=search) |
                Q(ingrediente__categoria__icontains=search) |
                Q(ingrediente__descripcion__icontains=search)
            )

        orden = self.request.GET.get("orden", "ingrediente__categoria")
        orden_validos = [
            "ingrediente__categoria", "-ingrediente__categoria",
            "ingrediente__nombre", "-ingrediente__nombre",
            "cantidad_actual", "-cantidad_actual",
            "costo_unitario", "-costo_unitario",
        ]
        if orden in orden_validos:
            qs = qs.order_by(orden)
        else:
            qs = qs.order_by("ingrediente__categoria", "ingrediente__nombre")

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_cliente()
        qs_ing = Ingrediente.objects.filter(cliente=cliente) if cliente else Ingrediente.objects.all()
        context["categorias"] = qs_ing.values_list("categoria", flat=True).exclude(categoria="").distinct().order_by("categoria")
        context["categoria_actual"] = self.request.GET.get("categoria", "")
        context["estado_actual"] = self.request.GET.get("estado", "")
        context["search"] = self.request.GET.get("search", "")
        context["orden_actual"] = self.request.GET.get("orden", "ingrediente__categoria")
        context["alertas"] = self.request.GET.get("alertas", "")
        context["total_ingredientes"] = self.get_queryset().count()
        context["bajo_stock_count"] = Inventario.objects.filter(
            Q(cantidad_actual__lte=F("ingrediente__stock_minimo"), ingrediente__stock_minimo__gt=0) |
            Q(cantidad_actual=0)
        ).count()
        return context


class IngredienteCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Ingrediente
    template_name = "inventario/ingrediente_form.html"
    fields = ["nombre", "descripcion", "categoria", "stock_minimo"]
    success_url = reverse_lazy("inventario:list")
    permission = "can_manage_inventario"

    def get_form(self, form_class=None):
        cliente = self.get_cliente()
        if cliente:
            Ingrediente.objects.filter(
                cliente=cliente,
                inventario__isnull=True
            ).annotate(
                has_compras=models.Exists(CompraItem.objects.filter(ingrediente=models.OuterRef('pk')))
            ).filter(has_compras=False).delete()
        form = super().get_form(form_class)
        return form

    def form_valid(self, form):
        cliente = self.get_cliente()
        if not cliente:
            cliente = getattr(self.request.user, 'cliente', None)
        if cliente:
            form.instance.cliente = cliente
        form.instance.nombre = form.instance.nombre.strip()
        try:
            response = super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, f"Ya existe un ingrediente con el nombre '{form.instance.nombre}'")
            return self.form_invalid(form)
        # Crear registro de inventario con stock 0
        Inventario.objects.get_or_create(
            ingrediente=form.instance,
            defaults={"cantidad_actual": 0, "unidad": Ingrediente.Unidad.UNIDAD, "costo_unitario": 0}
        )
        messages.success(self.request, f"Ingrediente '{form.instance.nombre}' creado")
        return response

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)


class InventarioUpdateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Inventario
    template_name = "inventario/form.html"
    fields = ["cantidad_actual", "unidad", "costo_unitario", "proveedor"]
    success_url = reverse_lazy("inventario:list")
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(ingrediente__cliente=cliente)
        return qs

    def form_valid(self, form):
        messages.success(self.request, "Inventario actualizado")
        return super().form_valid(form)


class InventarioDeleteView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Inventario
    success_url = reverse_lazy("inventario:list")
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(ingrediente__cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        ingrediente = self.object.ingrediente
        nombre = ingrediente.nombre
        try:
            MovimientoInventario.objects.filter(inventario=self.object).delete()
            CompraItem.objects.filter(ingrediente=ingrediente).delete()
            self.object.delete()
            ingrediente.delete()
            messages.success(request, f"Inventario de '{nombre}' eliminado permanentemente")
        except Exception as e:
            messages.error(request, f"No se pudo eliminar '{nombre}': {str(e)}")
        return redirect(self.success_url)


class MovimientoListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = MovimientoInventario
    template_name = "inventario/movimientos.html"
    context_object_name = "movimientos"
    paginate_by = 50
    permission = "can_view_inventario"

    def get_queryset(self):
        qs = super().get_queryset().select_related(
            "inventario__ingrediente",
            "usuario",
        )
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(inventario__ingrediente__cliente=cliente)
        return qs


@login_required
def registrar_movimiento(request, pk):
    if not request.user.has_perm("can_manage_inventario"):
        messages.error(request, "No tienes permiso")
        return redirect("core:dashboard")
    cliente = getattr(request.user, 'cliente', None)
    qs = Inventario.objects.select_related("ingrediente").all()
    if cliente:
        qs = qs.filter(ingrediente__cliente=cliente)
    inventario = get_object_or_404(qs, pk=pk)
    if request.method == "POST":
        tipo = request.POST.get("tipo")
        cantidad_str = request.POST.get("cantidad", "0")
        costo_str = request.POST.get("costo", "0")
        motivo = request.POST.get("motivo", "")

        validos = [t[0] for t in MovimientoInventario.Tipo.choices]
        if tipo not in validos:
            messages.error(request, "Tipo de movimiento no valido")
            return redirect("inventario:list")

        try:
            from decimal import Decimal
            cantidad = Decimal(cantidad_str)
            costo = Decimal(costo_str)
            if cantidad < 0:
                messages.error(request, "La cantidad no puede ser negativa")
                return redirect("inventario:list")
        except Exception:
            messages.error(request, "Cantidad no valida")
            return redirect("inventario:list")

        if tipo in (MovimientoInventario.Tipo.ENTRADA, MovimientoInventario.Tipo.DEVOLUCION):
            inventario.cantidad_actual += cantidad
            if costo > 0:
                inventario.costo_unitario = costo
        elif tipo in (MovimientoInventario.Tipo.SALIDA, MovimientoInventario.Tipo.MERMA):
            inventario.cantidad_actual = max(0, inventario.cantidad_actual - cantidad)
        elif tipo == MovimientoInventario.Tipo.AJUSTE:
            inventario.cantidad_actual = cantidad
            if costo > 0:
                inventario.costo_unitario = costo

        inventario.save()

        MovimientoInventario.objects.create(
            inventario=inventario,
            tipo=tipo,
            cantidad=cantidad,
            costo_unitario=costo,
            motivo=motivo,
            usuario=request.user,
        )

        if inventario.bajo_stock:
            messages.warning(request, f"Alerta! {inventario.ingrediente.nombre} tiene stock bajo")
        else:
            messages.success(request, "Movimiento registrado correctamente")

    return redirect("inventario:list")


class CompraItemForm(forms.ModelForm):
    class Meta:
        model = CompraItem
        fields = ["ingrediente", "cantidad", "costo_unitario"]
        widgets = {
            "ingrediente": forms.Select(attrs={"class": "compra-form-input"}),
            "cantidad": forms.NumberInput(attrs={"class": "compra-form-input", "step": "0.001"}),
            "costo_unitario": forms.NumberInput(attrs={"class": "compra-form-input", "step": "0.01"}),
        }

    def __init__(self, *args, **kwargs):
        self._cliente = kwargs.pop("cliente", None)
        super().__init__(*args, **kwargs)
        qs = Ingrediente.objects.filter(inventario__isnull=False).distinct().order_by("categoria", "nombre")
        if self._cliente:
            qs = qs.filter(cliente=self._cliente)
        self.fields["ingrediente"].widget = forms.Select(
            choices=[("", "---------")] + [
                (p.pk, f"{p.nombre}{' [' + p.categoria + ']' if p.categoria else ''}") for p in qs
            ],
            attrs={"class": "compra-form-input"},
        )
        self.fields["cantidad"].widget.attrs.update({"class": "compra-form-input", "step": "0.001"})
        self.fields["costo_unitario"].widget.attrs.update({"class": "compra-form-input", "step": "0.01"})
        for field in self.fields.values():
            field.required = False

    def clean(self):
        cleaned = super().clean()
        ing = cleaned.get('ingrediente')
        cant = cleaned.get('cantidad')
        costo = cleaned.get('costo_unitario')
        if not ing and not cant and not costo:
            return cleaned
        if not ing:
            self.add_error('ingrediente', 'Este campo es obligatorio')
        if not cant:
            self.add_error('cantidad', 'Este campo es obligatorio')
        if not costo:
            self.add_error('costo_unitario', 'Este campo es obligatorio')
        if ing:
            cliente = getattr(self, '_cliente', None)
            if cliente and ing.cliente_id != cliente.pk:
                self.add_error('ingrediente', 'Este ingrediente no pertenece a tu restaurante')
        return cleaned


CompraItemFormSet = forms.modelformset_factory(
    CompraItem,
    form=CompraItemForm,
    extra=5,
    can_delete=True,
)


class BaseCompraItemFormSet(forms.BaseModelFormSet):
    def clean(self):
        super().clean()
        valid = []
        for form in self.forms:
            if hasattr(form, 'cleaned_data') and form.cleaned_data:
                if form.cleaned_data.get('DELETE'):
                    continue
                if form.cleaned_data.get('ingrediente'):
                    valid.append(form)
        if not valid:
            raise forms.ValidationError("Debes agregar al menos un ingrediente")


def get_compra_item_formset(**kwargs):
    return forms.modelformset_factory(
        CompraItem,
        form=CompraItemForm,
        extra=10,
        can_delete=True,
        formset=BaseCompraItemFormSet,
    )


class CompraListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Compra
    template_name = "inventario/compras.html"
    context_object_name = "compras"
    paginate_by = 20
    permission = "can_manage_inventario"


class CompraCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Compra
    template_name = "inventario/compra_form.html"
    fields = ["proveedor_fk", "proveedor", "notas"]
    success_url = reverse_lazy("inventario:compras")
    permission = "can_manage_inventario"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        cliente = self.get_cliente()
        proveedores = Proveedor.objects.filter(cliente=cliente, activo=True) if cliente else Proveedor.objects.none()
        form.fields["proveedor_fk"] = forms.ModelChoiceField(
            queryset=proveedores, required=False,
            widget=forms.Select(attrs={"class": "md3-input", "onchange": "document.getElementById('id_proveedor').value = this.options[this.selectedIndex].text"}),
            label="Proveedor (seleccionar)",
        )
        form.fields["proveedor"].widget = forms.TextInput(attrs={
            "class": "md3-input",
            "placeholder": " ",
        })
        form.fields["notas"].widget = forms.Textarea(attrs={
            "class": "md3-input",
            "rows": 2,
            "placeholder": " ",
            "style": "resize:vertical; min-height:56px;",
        })
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_cliente()
        ingredientes = Ingrediente.objects.filter(activo=True).order_by("categoria", "nombre")
        if cliente:
            ingredientes = ingredientes.filter(cliente=cliente)

        inv_qs = Inventario.objects.all()
        if cliente:
            inv_qs = inv_qs.filter(ingrediente__cliente=cliente)

        inv_map = {}
        for inv in inv_qs.select_related("ingrediente"):
            inv_map[inv.ingrediente_id] = {
                "costo": float(inv.costo_unitario) if inv.costo_unitario else 0,
                "unidad": inv.get_unidad_display(),
                "stock": float(inv.cantidad_actual) if inv.cantidad_actual else 0,
            }

        context["ingredientes_json"] = [
            {
                "id": ing.pk,
                "nombre": ing.nombre,
                "categoria": ing.categoria or "",
                "unidad": inv_map.get(ing.pk, {}).get("unidad", ""),
                "costo_sugerido": inv_map.get(ing.pk, {}).get("costo", 0),
                "stock_actual": inv_map.get(ing.pk, {}).get("stock", 0),
            }
            for ing in ingredientes
        ]
        return context

    def form_valid(self, form):
        items_data = []
        idx = 0
        while True:
            ing_id = self.request.POST.get(f"items[{idx}][ingrediente]")
            if not ing_id:
                break
            cantidad = self.request.POST.get(f"items[{idx}][cantidad]")
            costo = self.request.POST.get(f"items[{idx}][costo_unitario]")
            if ing_id and cantidad and costo:
                try:
                    ing = Ingrediente.objects.get(pk=ing_id)
                    c = float(cantidad)
                    co = float(costo)
                    if c > 0 and co >= 0:
                        cliente = self.get_cliente()
                        if cliente and ing.cliente_id != cliente.pk:
                            raise forms.ValidationError(f"Ingrediente '{ing.nombre}' no pertenece a tu restaurante")
                        items_data.append((ing, c, co))
                except (Ingrediente.DoesNotExist, ValueError):
                    pass
            idx += 1

        if not items_data:
            messages.error(self.request, "Debes agregar al menos un ingrediente")
            return self.form_invalid(form)

        cliente = self.get_cliente()
        form.instance.usuario = self.request.user
        form.instance.estado = Compra.Estado.PENDIENTE
        if not form.instance.cliente_id and cliente:
            form.instance.cliente = cliente
        response = super().form_valid(form)

        for ing, cant, costo in items_data:
            CompraItem.objects.create(
                compra=self.object,
                ingrediente=ing,
                cantidad=cant,
                costo_unitario=costo,
            )

        messages.success(self.request, f"Compra {self.object.folio} registrada como pendiente")
        return response


@require_GET
def ingredientes_api(request):
    cliente = getattr(request.user, "cliente", None) if request.user.is_authenticated else None
    qs = Ingrediente.objects.filter(activo=True).order_by("categoria", "nombre")
    if cliente:
        qs = qs.filter(cliente=cliente)

    inv_qs = Inventario.objects.all()
    if cliente:
        inv_qs = inv_qs.filter(ingrediente__cliente=cliente)

    inv_map = {}
    for inv in inv_qs.select_related("ingrediente"):
        inv_map[inv.ingrediente_id] = {
            "costo": float(inv.costo_unitario) if inv.costo_unitario else 0,
            "unidad": inv.get_unidad_display(),
            "stock": float(inv.cantidad_actual) if inv.cantidad_actual else 0,
        }

    data = [
        {
            "id": ing.pk,
            "nombre": ing.nombre,
            "categoria": ing.categoria or "",
            "unidad": inv_map.get(ing.pk, {}).get("unidad", ""),
            "costo_sugerido": inv_map.get(ing.pk, {}).get("costo", 0),
            "stock_actual": inv_map.get(ing.pk, {}).get("stock", 0),
        }
        for ing in qs
    ]
    return JsonResponse(data, safe=False)


class CompraDetailView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Compra
    template_name = "inventario/compra_detalle.html"
    context_object_name = "compra"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs


class CompraRecibirView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Compra
    template_name = "inventario/compra_detalle.html"
    context_object_name = "compra"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        compra = self.get_object()
        if compra.estado != Compra.Estado.PENDIENTE:
            messages.error(request, "Esta compra ya fue procesada")
            return redirect("inventario:compras")

        cliente = self.get_cliente()

        for item in compra.items.all():
            inv_qs = Inventario.objects.all()
            if cliente:
                inv_qs = inv_qs.filter(ingrediente__cliente=cliente)
            inv, created = inv_qs.get_or_create(
                ingrediente=item.ingrediente,
                defaults={
                    "cantidad_actual": item.cantidad,
                    "unidad": Ingrediente.Unidad.UNIDAD,
                    "costo_unitario": item.costo_unitario,
                    "proveedor": compra.proveedor,
                }
            )
            if not created:
                inv.cantidad_actual += item.cantidad
                inv.costo_unitario = item.costo_unitario
                inv.proveedor = compra.proveedor
                inv.save()

            MovimientoInventario.objects.create(
                inventario=inv,
                tipo=MovimientoInventario.Tipo.ENTRADA,
                cantidad=item.cantidad,
                costo_unitario=item.costo_unitario,
                motivo=f"Compra {compra.folio}",
                usuario=request.user,
            )

        compra.estado = Compra.Estado.RECIBIDA
        compra.save()

        config = ConfigRestaurante.get_config(cliente)
        dias = config.dias_credito_proveedor

        CuentaPorPagar.objects.create(
            cliente=cliente,
            proveedor=compra.proveedor,
            proveedor_fk=compra.proveedor_fk,
            compra=compra,
            monto_total=compra.total,
            fecha_vencimiento=timezone.now().date() + timezone.timedelta(days=dias),
            usuario=request.user,
            estado=CuentaPorPagar.Estado.PENDIENTE,
        )

        messages.success(request, f"Compra {compra.folio} recibida - Inventario actualizado - Cuenta por pagar generada")
        return redirect("inventario:compras")


class CompraCancelarView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Compra
    template_name = "inventario/compra_detalle.html"
    context_object_name = "compra"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        compra = self.get_object()
        if compra.estado == Compra.Estado.RECIBIDA:
            messages.error(request, "No puedes cancelar una compra ya recibida")
            return redirect("inventario:compras")
        compra.estado = Compra.Estado.CANCELADA
        compra.save()
        messages.success(request, f"Compra {compra.folio} cancelada")
        return redirect("inventario:compras")


class CompraDeleteView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Compra
    success_url = reverse_lazy("inventario:compras")
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            folio = self.object.folio
            self.object.delete()
            messages.success(request, f"Compra {folio} eliminada")
        except Exception:
            messages.error(request, "No se puede eliminar: tiene datos asociados")
        return redirect(self.success_url)


class CuentaPorPagarListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = CuentaPorPagar
    template_name = "inventario/cuentas_list.html"
    context_object_name = "cuentas"
    paginate_by = 20
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset().select_related("compra")
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)

        estado = self.request.GET.get("estado")
        if estado:
            qs = qs.filter(estado=estado)

        proveedor = self.request.GET.get("proveedor")
        if proveedor:
            qs = qs.filter(proveedor__icontains=proveedor)

        search = self.request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(proveedor__icontains=search) |
                Q(folio__icontains=search) |
                Q(notas__icontains=search)
            )

        fecha_vencimiento = self.request.GET.get("fecha_vencimiento")
        if fecha_vencimiento:
            qs = qs.filter(fecha_vencimiento__lte=fecha_vencimiento)

        orden = self.request.GET.get("orden", "-fecha_creacion")
        orden_validos = [
            "fecha_creacion", "-fecha_creacion",
            "fecha_vencimiento", "-fecha_vencimiento",
            "monto_total", "-monto_total",
            "proveedor", "-proveedor",
        ]
        if orden in orden_validos:
            qs = qs.order_by(orden)
        elif orden in ("monto_pendiente", "-monto_pendiente"):
            pendiente = ExpressionWrapper(
                F("monto_total") - F("monto_pagado"),
                output_field=DecimalField()
            )
            qs = qs.annotate(monto_pendiente_calc=pendiente)
            prefix = "-" if orden.startswith("-") else ""
            qs = qs.order_by(f"{prefix}monto_pendiente_calc")
        else:
            qs = qs.order_by("-fecha_creacion")

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_cliente()
        cuentas_qs = CuentaPorPagar.objects.all()
        if cliente:
            cuentas_qs = cuentas_qs.filter(cliente=cliente)
        context["total_pendiente"] = cuentas_qs.filter(
            estado__in=["pendiente", "parcial", "vencida"]
        ).aggregate(total=Sum(F("monto_total") - F("monto_pagado")))["total"] or 0
        context["cuentas_vencidas"] = cuentas_qs.filter(estado="vencida").count()
        context["estado_actual"] = self.request.GET.get("estado", "")
        context["proveedor_actual"] = self.request.GET.get("proveedor", "")
        context["search"] = self.request.GET.get("search", "")
        context["fecha_vencimiento"] = self.request.GET.get("fecha_vencimiento", "")
        context["orden_actual"] = self.request.GET.get("orden", "-fecha_creacion")
        return context


class CuentaPorPagarCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = CuentaPorPagar
    template_name = "inventario/cuenta_form.html"
    fields = ["proveedor_fk", "proveedor", "monto_total", "fecha_vencimiento", "notas"]
    success_url = reverse_lazy("inventario:cuentas")
    permission = "can_manage_inventario"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        cliente = self.get_cliente()
        proveedores = Proveedor.objects.filter(cliente=cliente, activo=True) if cliente else Proveedor.objects.none()
        form.fields["proveedor_fk"] = forms.ModelChoiceField(
            queryset=proveedores, required=False,
            widget=forms.Select(attrs={"class": "md3-input", "onchange": "document.getElementById('id_proveedor').value = this.options[this.selectedIndex].text"}),
            label="Proveedor (seleccionar)",
        )
        form.fields["proveedor"].widget = forms.TextInput(attrs={
            "class": "md3-input",
            "placeholder": " ",
        })
        form.fields["proveedor"].help_text = "O escribe el nombre manualmente si no está en la lista."
        return form

    def form_valid(self, form):
        form.instance.usuario = self.request.user
        form.instance.monto_pagado = 0
        form.instance.estado = CuentaPorPagar.Estado.PENDIENTE
        response = super().form_valid(form)
        messages.success(self.request, f"Cuenta {form.instance.folio} creada")
        return response


class CuentaPorPagarDetailView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = CuentaPorPagar
    template_name = "inventario/cuenta_detalle.html"
    context_object_name = "cuenta"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs


class CuentaPorPagarPagoView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = CuentaPorPagar
    template_name = "inventario/cuenta_detalle.html"
    context_object_name = "cuenta"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        cuenta = self.get_object()
        if cuenta.estado in (CuentaPorPagar.Estado.PAGADA, CuentaPorPagar.Estado.CANCELADA):
            messages.error(request, "Esta cuenta ya fue procesada")
            return redirect("inventario:cuentas")
        monto_str = request.POST.get("monto", "")
        try:
            from decimal import Decimal
            monto = Decimal(monto_str)
            if monto <= 0:
                messages.error(request, "El monto debe ser mayor a 0")
                return redirect("inventario:cuenta_detalle", pk=kwargs['pk'])
            if monto > cuenta.monto_pendiente:
                messages.error(request, f"El monto no puede ser mayor a {cuenta.monto_pendiente}")
                return redirect("inventario:cuenta_detalle", pk=kwargs['pk'])
        except Exception:
            messages.error(request, "Monto no valido")
            return redirect("inventario:cuenta_detalle", pk=kwargs['pk'])
        cuenta.registrar_pago(monto, request.user)
        messages.success(request, f"Pago registrado - {cuenta.monto_pendiente} pendiente")
        return redirect("inventario:cuenta_detalle", pk=kwargs['pk'])


class CuentaPorPagarCancelarView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = CuentaPorPagar
    template_name = "inventario/cuenta_detalle.html"
    context_object_name = "cuenta"
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        cuenta = self.get_object()
        if cuenta.estado == CuentaPorPagar.Estado.PAGADA:
            messages.error(request, "No puedes cancelar una cuenta ya pagada")
            return redirect("inventario:cuentas")
        cuenta.estado = CuentaPorPagar.Estado.CANCELADA
        cuenta.save()
        messages.success(request, f"Cuenta {cuenta.folio} cancelada")
        return redirect("inventario:cuentas")


class CuentaPorPagarDeleteView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = CuentaPorPagar
    success_url = reverse_lazy("inventario:cuentas")
    permission = "can_manage_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            folio = self.object.folio
            self.object.delete()
            messages.success(request, f"Cuenta {folio} eliminada")
        except Exception:
            messages.error(request, "No se puede eliminar: tiene datos asociados")
        return redirect(self.success_url)


def inventario_pdf(request):
    from django.http import HttpResponse
    from core.pdf_utils import generate_inventario_pdf

    cliente = getattr(request.user, "cliente", None)
    config = ConfigRestaurante.get_config(cliente)
    qs = Inventario.objects.select_related("ingrediente").order_by("ingrediente__categoria", "ingrediente__nombre")
    if cliente:
        qs = qs.filter(ingrediente__cliente=cliente)

    buffer = generate_inventario_pdf(qs, config.nombre, config.simbolo_moneda)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="inventario.pdf"'
    return response


def compras_pdf(request):
    from django.http import HttpResponse
    from core.pdf_utils import generate_compras_pdf

    cliente = getattr(request.user, "cliente", None)
    config = ConfigRestaurante.get_config(cliente)
    qs = Compra.objects.all().order_by("-fecha")
    if cliente:
        qs = qs.filter(cliente=cliente)

    buffer = generate_compras_pdf(qs, config.nombre, config.simbolo_moneda)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="compras.pdf"'
    return response


# ─── Proveedores CRUD ───────────────────────────────────────────────

class ProveedorListView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Proveedor
    template_name = "inventario/proveedor_list.html"
    context_object_name = "proveedores"
    permission = "can_view_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs.filter(activo=True)


class ProveedorCreateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Proveedor
    template_name = "inventario/proveedor_form.html"
    fields = ["nombre", "contacto", "telefono", "email", "direccion", "notas"]
    permission = "can_view_inventario"

    def form_valid(self, form):
        form.instance.cliente = self.get_cliente()
        messages.success(self.request, "Proveedor creado exitosamente.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy("inventario:proveedores")


class ProveedorUpdateView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Proveedor
    template_name = "inventario/proveedor_form.html"
    fields = ["nombre", "contacto", "telefono", "email", "direccion", "notas"]
    permission = "can_view_inventario"

    def get_success_url(self):
        return reverse_lazy("inventario:proveedores")


class ProveedorDeleteView(ClienteScopeMixin, PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Proveedor
    template_name = "inventario/proveedor_confirm_delete.html"
    permission = "can_view_inventario"

    def get_queryset(self):
        qs = super().get_queryset()
        cliente = self.get_cliente()
        if cliente:
            qs = qs.filter(cliente=cliente)
        return qs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.activo = False
        self.object.save()
        messages.success(request, "Proveedor desactivado exitosamente.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy("inventario:proveedores")


# ─── Export Excel ───────────────────────────────────────────────────

@login_required
def inventario_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    cliente = ClienteScopeMixin.get_cliente_static(request)
    qs = Inventario.objects.filter(ingrediente__cliente=cliente).select_related("ingrediente")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_font = Font(name="Arial", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Ingrediente", "Categoria", "Cantidad", "Unidad", "Costo Unitario", "Valor Total", "Stock Minimo", "Proveedor"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, inv in enumerate(qs, 2):
        data = [
            inv.ingrediente.nombre,
            inv.ingrediente.categoria,
            float(inv.cantidad_actual),
            inv.get_unidad_display(),
            float(inv.costo_unitario),
            float(inv.valor_total),
            float(inv.ingrediente.stock_minimo),
            inv.proveedor,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 20

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response


@login_required
def compras_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    cliente = ClienteScopeMixin.get_cliente_static(request)
    qs = Compra.objects.filter(cliente=cliente).prefetch_related("items__ingrediente")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    header_font = Font(name="Arial", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Folio", "Proveedor", "Fecha", "Estado", "Total", "Notas"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, compra in enumerate(qs, 2):
        data = [
            compra.folio,
            compra.proveedor,
            compra.fecha.strftime("%d/%m/%Y %H:%M"),
            compra.get_estado_display(),
            float(compra.total),
            compra.notas,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="compras.xlsx"'
    wb.save(response)
    return response
